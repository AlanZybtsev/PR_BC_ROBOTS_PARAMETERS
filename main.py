import os
import pandas as pd
import datetime
from openpyxl import load_workbook
from pandas import ExcelWriter

# ДАТА НА МОМЕНТ ПРОГОНА СКРИПТА
now = datetime.datetime.now().date().strftime("%d.%m.%Y")

# СПИСОК МОДЕЛЕЙ КАМАЗ
models_list = ['AXOR', 'F5320', 'F5321', 'F5410', 'F5460', 'F6460', 'F6461',\
               'F53205', 'F54105', 'F535003', 'F535004', 'KOMPAS']

# СПИСОК НОМЕРОВ ЦИРКУЛЯЦИОННЫХ СИСТЕМ
# С01 - RAL2009
# C01 - RAL9069
# C03 - RAL1033
# C04 - RAL5010
# C05 - RAL9010
# C06 - RAL1021
# C07 - RAL3004
# C08 - RAL3000(КНБ)
circ_sys_qty = [f'C0{i}' for i in range(1, 9)]

# СПИСОК РОБОТОВ ВС1
ROB_PAR_dir = ".\\ROB_PAR"
BC1_dirs = [i for i in os.listdir(ROB_PAR_dir)]

# ОТФИЛЬТРОВАННЫЙ СЛОВАРЬ МОДЕЛЕЙ ДЛЯ КАЖДОГО РОБОТА
robot_model = {}
for dir in BC1_dirs:
    directory = f".\\ROB_PAR\\{dir}\\BRUSHES.INI"
    files = os.listdir(directory)
    fil_files = []
    for file in files:
        splited_file = file.split('.')
        if splited_file[0] in models_list and splited_file[1] in circ_sys_qty:
            fil_files.append(file)
    robot_model[dir]=fil_files

# СОЗДАНИЕ ПАПОК ДЛЯ КАЖОДОЙ МОДЕЛИ, !!!ЕСЛИ!!! НЕ СУЩЕСТВУЕТ
for model in models_list:
    for circ_sys in circ_sys_qty:
        os.makedirs(f'.\\PARAMETERS\\{model}', exist_ok=True)

# СПИСОК НОМЕРОВ РОБОТОВ
robot_numbers = list(robot_model.keys())

# НОМЕР КОЛОНКИ ДЛЯ ЗАПИСИ ДАТА ФРЕЙМОВ В ЭКСЕЛЬ ФАЙЛ
start_col = 0


for robot, file_names in robot_model.items():
    if robot != 'R11':
        # ЕСЛИ НОМЕР РОБОТА НЕ РАВЕН R11 КОЛОНКА ДЛЯ ЗАПИСИ ФРЕЙМА += 5
        start_col += 5

    for file_name in file_names:
        # ПУТЬ К НЕОБХОДИМОМУ ФАЙЛУ
        param_dir = f".\\ROB_PAR\\{robot}\\BRUSHES.INI\\{file_name}"
        brushes = {}

        with open(param_dir, 'r', encoding = 'utf-8') as f:
            brush_num = []
            for line in f:
                if line.startswith('[B'):
                    br = line.strip('[]\n')
                    br_num = br.replace(br[0:5], '')
                    brush_num.append(br_num)
            # ДОБАВЛЕНИЕ В СЛОВАРЬ brushes НОМЕРОВ КИСТЕЙ РОБОТА
            brushes[robot] = brush_num

        with open(param_dir, 'r', encoding = 'utf-8') as f:
            brush_f_r = []
            for line in f:
                if line.startswith('FM2'):
                    f_r = line[6:].strip()
                    brush_f_r.append(f_r)
            # ДОБАВЛЕНИЕ В СЛОВАРЬ brushes РАСХОДА ЛКМ ДЛЯ КАЖДОЙ КИСТИ
            brushes['FR'] = brush_f_r

        with open(param_dir, 'r', encoding = 'utf-8') as f:
            brush_air = []
            for line in f:
                if line.startswith('LM1'):
                    air = line[6:].strip()
                    brush_air.append(air)
            # ДОБАВЛЕНИЕ В СЛОВАРЬ РАСХОДА ВОЗДУХА ДЛЯ КАЖДОЙ КИСТИ
            brushes['AIR'] = brush_air

        with open(param_dir, 'r', encoding = 'utf-8') as f:
            brush_rot = []
            for line in f:
                if line.startswith('LM2'):
                    rot = line[6:].strip()
                    brush_rot.append(rot)
            # ДОБАВЛЕНИЕ В СЛОВАРЬ brushes ОБОРОТОВ ЧАШКИ ДЛЯ КАЖДОЙ КИСТИ
            brushes['ROT'] = brush_rot

        with open(param_dir, 'r', encoding = 'utf-8') as f:
            brush_amp = []
            for line in f:
                if line.startswith('HT1'):
                    amp = line[6:].strip()
                    brush_amp.append(amp)
            # ДОБАВЛЕНИЕ В СЛОВАРЬ brushes ТОКА ДЛЯ КАЖДОЙ КИСТИ
            brushes['HT'] = brush_amp

        # УДАЛЕНИЕ ИЗ СЛОВАРЯ НОМЕРОВ КИСТЕЙ И ИХ ПАРАМЕТРОВ, ЕСЛИ РАСХОД ЛКМ = 0
        zero_index = [index for index, number in enumerate(brushes['FR']) if number == '0']
        n = 0
        for z_i in zero_index:
            brushes[robot].pop(z_i - n)
            brushes['FR'].pop(z_i - n)
            brushes['AIR'].pop(z_i - n)
            brushes['ROT'].pop(z_i - n)
            brushes['HT'].pop(z_i - n)
            n += 1

        # model - МОДЕЛЬ, color - ЦВЕТ, destiny_file_dir - ПУТЬ К СОЗДАНИЮ ФАЙЛА, params - ФРЕЙМ
        model = file_name.split('.')[0]
        color = file_name.split('.')[1]
        destiny_file_dir = f".\\PARAMETERS\\{model}\\{model} - {now}.xlsx"
        params = pd.DataFrame(brushes)

        # СПИСОК ГОТОВЫХ ФАЙЛОВ В КОНЕЧНОЙ ДИРЕКТОРИИ
        destiny_dir_files = os.listdir(f".\\PARAMETERS\\{model}")

        # ЕСЛИ СПИСОК destiny_dir_files ПУСТ, СОЗДАНИЕ ФАЙЛА И ЗАПИСЬ ПЕРВОГО ФАЙЛА
        if len(destiny_dir_files) == 0:
            params.to_excel(destiny_file_dir, sheet_name = color, index=False, startcol=start_col, startrow=1)

        # ЕСЛИ СПИСОК НЕ ПУСТ ДОБАВЛЕНИЕ НОВЫХ ЛИСТОВ С ПАРАМЕТРАМИ НАНЕСЕНИЯ
        else:
            with ExcelWriter(destiny_file_dir, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                params.to_excel(writer, sheet_name = color, index=False, startcol=start_col, startrow=1)

            # ЗАПИСЬ НА КАЖДЫЙ ЛИСТ МОДЕЛИ, ЦВЕТА И ДАТЫ ЗАПИСИ
            wb = load_workbook(destiny_file_dir)
            ws = wb[color]
            model_coord = "A1"
            ws[model_coord].value = model
            color_coord = "B1"
            ws[color_coord].value = color
            date_coord = "C1"
            ws[date_coord].value = now
            wb.save(destiny_file_dir)
            wb.close()
